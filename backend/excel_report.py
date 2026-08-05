"""
excel_report.py - QA Report Excel Generator
สร้าง Excel Report จากผลการวิเคราะห์ QA ด้วย Gemini AI
"""

import os
import re
import json
import logging
import uuid
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# Reports output directory
REPORTS_DIR = Path(__file__).parent.parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")


def parse_qa_report_with_ai(report_text: str, filename: str) -> list[dict]:
    """
    Use Gemini AI to parse the QA report text into structured rows for Excel.
    Falls back to regex parsing if AI fails.
    """
    try:
        from ocr_engine import _get_gemini_client
        client = _get_gemini_client()
        gemini_model = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash')

        prompt = f"""คุณคือผู้เชี่ยวชาญด้านการวิเคราะห์รายงาน QA (Quality Assurance)
กรุณาแปลงรายงานการตรวจสอบด้านล่างนี้ ให้เป็น JSON Array ที่แต่ละ Object มี Key ดังนี้:

- "document_page": ชื่อเอกสาร / หน้าที่เกี่ยวข้อง (ถ้าไม่ระบุให้ใส่ชื่อไฟล์ "{filename}")
- "issue": ประเด็นที่พบ (สรุปสั้นๆ)
- "severity": ระดับความรุนแรง (Critical / High / Medium / Low / Info)
- "check_type": ประเภทการตรวจ (เช่น SRS Completeness, UI/UX, Security, Performance)
- "found_incorrect": ข้อความในเอกสาร (สิ่งที่พบเจอและไม่ถูกต้อง)
- "correct_value": สิ่งที่ควรเป็น (ค่าที่ควรจะเป็น หรือแนวทางที่ถูก ถ้าไม่มีให้ใส่ "-")
- "recommendation": ข้อเสนอแนะ (คำแนะนำในการแก้ไข)

**สำคัญมาก**: ตอบเป็น JSON Array เท่านั้น ห้ามมี text อื่นนอกจาก JSON
ถ้าไม่มีประเด็นใดๆ ให้ตอบ []

=== รายงาน QA ===
{report_text}
"""

        fallback_models = [gemini_model, 'gemini-2.5-flash', 'gemini-2.5-flash-lite']
        gemini_res = None
        
        for current_model in fallback_models:
            for attempt in range(2):
                try:
                    gemini_res = client.models.generate_content(
                        model=current_model,
                        contents=prompt,
                    )
                    break
                except Exception as e:
                    error_msg = str(e)
                    if '429' in error_msg or 'RESOURCE_EXHAUSTED' in error_msg:
                        import time
                        time.sleep(5)
                        continue
                    else:
                        break
            if gemini_res:
                break

        if gemini_res and gemini_res.text:
            # Extract JSON from response
            text = gemini_res.text.strip()
            # Remove markdown code block if present
            if text.startswith("```"):
                text = re.sub(r'^```(?:json)?\s*', '', text)
                text = re.sub(r'\s*```$', '', text)
            
            rows = json.loads(text)
            if isinstance(rows, list):
                return rows
    
    except Exception as e:
        logger.error(f"AI parsing failed, falling back to regex: {e}")

    # Fallback: basic regex parsing
    return parse_qa_report_regex(report_text, filename)


def parse_qa_report_regex(report_text: str, filename: str) -> list[dict]:
    """
    Fallback regex-based parser for QA report text.
    Tries to extract bullet points and numbered items as issues.
    """
    rows = []
    lines = report_text.split('\n')
    current_section = ""
    item_count = 0
    
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        
        # Detect section headers (## or ### or bold **)
        if stripped.startswith('#'):
            current_section = re.sub(r'^#+\s*', '', stripped).strip()
            continue
        if stripped.startswith('**') and stripped.endswith('**'):
            current_section = stripped.strip('*').strip()
            continue
        
        # Detect bullet points or numbered items as issues
        issue_match = re.match(r'^[\-\*\•]\s+(.+)', stripped)
        num_match = re.match(r'^\d+[\.\)]\s+(.+)', stripped)
        
        if issue_match or num_match:
            content = (issue_match or num_match).group(1)
            item_count += 1
            
            # Determine severity based on keywords
            severity = "Low"
            if any(kw in content.lower() for kw in ['critical', 'ร้ายแรง', 'วิกฤต']):
                severity = "Critical"
            elif any(kw in content.lower() for kw in ['high', 'major', 'สำคัญ', 'ผิดพลาด', 'ขัดแย้ง']):
                severity = "High"
            elif any(kw in content.lower() for kw in ['medium', 'ปานกลาง']):
                severity = "Medium"
            elif any(kw in content.lower() for kw in ['ถูกต้อง', 'สอดคล้อง', 'ผ่าน', 'info']):
                severity = "Info"
            
            # Determine check type from section
            check_type = "ทั่วไป"
            if any(kw in current_section for kw in ['สอดคล้อง', 'Conformity', 'SRS']):
                check_type = "SRS Completeness"
            elif any(kw in current_section for kw in ['ขัดแย้ง', 'ผิดพลาด', 'Error']):
                check_type = "ข้อมูลขัดแย้ง"
            
            rows.append({
                "document_page": filename,
                "issue": content[:100],
                "severity": severity,
                "check_type": check_type,
                "found_incorrect": content,
                "correct_value": "-",
                "recommendation": "-"
            })
    
    # If nothing was parsed, create a single summary row
    if not rows:
        rows.append({
            "document_page": filename,
            "issue": "รายงานการตรวจสอบ QA",
            "severity": "Info",
            "check_type": "สรุปผล",
            "found_incorrect": report_text[:500],
            "correct_value": "-",
            "recommendation": "ดูรายละเอียดในรายงานฉบับเต็ม"
        })
    
    return rows


def generate_qa_excel(
    report_text: str,
    filename: str,
    doc_type: str = "",
    project_code: str = "",
    group_name: str = "",
    group_type: str = "",
    transaction_id: str = None
) -> str:
    """
    Generate a professional QA Report Excel file.
    Returns the file path of the generated Excel.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    # Parse report into structured rows
    rows = parse_qa_report_with_ai(report_text, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Report"
    
    # === Styles ===
    # Header style
    header_font = Font(name='Tahoma', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1A4441', end_color='1A4441', fill_type='solid') # Dark Teal
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Title style
    title_font = Font(name='Tahoma', bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid') # Dark Blue
    
    subtitle_font = Font(name='Tahoma', bold=False, italic=True, size=11, color='333333')
    subtitle_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') # Light Gray
    
    # Data styles
    data_font = Font(name='Tahoma', size=10)
    data_alignment = Alignment(vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Severity colors (Match screenshot)
    severity_fills = {
        'Critical': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'), # Red
        'High': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'), # Orange
        'Medium': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'), # Yellow
        'Low': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'Info': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    }
    severity_fonts = {
        'Critical': Font(name='Tahoma', bold=True, size=10, color='FFFFFF'),
        'High': Font(name='Tahoma', bold=True, size=10, color='FFFFFF'),
        'Medium': Font(name='Tahoma', bold=True, size=10, color='000000'),
        'Low': Font(name='Tahoma', bold=False, size=10, color='000000'),
        'Info': Font(name='Tahoma', bold=False, size=10, color='000000'),
    }
    
    # Border
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # === Title Section ===
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "QA Report Part 2 – Technical / Mockup vs Requirements / SRS Completeness"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:H2')
    subtitle_cell = ws['A2']
    project_str = project_code if project_code else (group_name if group_name else "2Hand To You (ระบบซื้อ-ขายสินค้ามือสอง)")
    subtitle_cell.value = f"โครงการ: {project_str} - Software Engineering, Burapha University, Team 7 | ผู้ตรวจสอบ: QA Auditor | วันที่ตรวจสอบ: {datetime.now().strftime('%d/%m/%Y')}"
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    
    # === Data Table ===
    table_start_row = 3
    
    # Column headers
    headers = [
        "ลำดับ",
        "เอกสาร / หน้า",
        "ประเด็นที่พบ",
        "ระดับความรุนแรง",
        "ประเภทการตรวจ",
        "ข้อความในเอกสาร",
        "สิ่งที่ควรเป็น",
        "ข้อเสนอแนะ"
    ]
    
    # Column widths
    col_widths = [8, 25, 40, 15, 25, 45, 45, 45]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[table_start_row].height = 30
    
    # Data rows
    for row_idx, row_data in enumerate(rows, 1):
        excel_row = table_start_row + row_idx
        
        severity = row_data.get('severity', 'Info')
        
        values = [
            row_idx,
            row_data.get('document_page', filename),
            row_data.get('issue', '-'),
            severity,
            row_data.get('check_type', '-'),
            row_data.get('found_incorrect', '-'),
            row_data.get('correct_value', '-'),
            row_data.get('recommendation', '-')
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment if col_idx not in [1, 4] else center_alignment
            cell.border = thin_border
            
            # Apply severity color
            if col_idx == 4:  # ระดับความรุนแรง
                cell.fill = severity_fills.get(severity, PatternFill())
                cell.font = severity_fonts.get(severity, data_font)
        
        ws.row_dimensions[excel_row].height = 50
    
    # === Footer ===
    footer_row = table_start_row + len(rows) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer_cell = ws.cell(row=footer_row, column=1, value="สร้างโดย Spectra QA Intelligent Analysis System | Powered by Gemini AI")
    footer_cell.font = Font(name='Tahoma', italic=True, size=9, color='999999')
    footer_cell.alignment = Alignment(horizontal='center')
    
    # === Save file ===
    report_id = transaction_id or str(uuid.uuid4())[:8]
    safe_filename = re.sub(r'[^\w\-.]', '_', filename.rsplit('.', 1)[0])
    excel_filename = f"QA_Report_{safe_filename}_{report_id}.xlsx"
    
    filepath = REPORTS_DIR / excel_filename
    wb.save(str(filepath))
    
    logger.info(f"Excel QA Report saved: {filepath}")
    return str(filepath)
