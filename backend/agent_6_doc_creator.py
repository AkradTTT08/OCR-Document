import json
import logging
import os
import google.generativeai as genai
from db_ingestion import get_db_connection

logger = logging.getLogger(__name__)

def create_qa_document(project_id: str, doc_type: str, doc_name: str, skill_id: int):
    """
    Agent 6: QA Document Creator
    Uses project requirements (Knowledge) and a selected Skill to generate a QA document.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Fetch Project Info
        cursor.execute("SELECT project_code FROM projects WHERE id = %s::uuid", (project_id,))
        project_res = cursor.fetchone()
        project_code = project_res[0] if project_res else "Unknown Project"

        # 2. Fetch Knowledge (Phase 1 Requirements)
        cursor.execute("""
            SELECT req_code, title, description, steps, expected_results
            FROM structured_requirements
            WHERE project_id = %s::uuid
        """, (project_id,))
        reqs = cursor.fetchall()
        
        formatted_reqs = []
        for req in reqs:
            formatted_reqs.append({
                "req_code": req[0],
                "title": req[1],
                "description": req[2],
                "steps": req[3],
                "expected_results": req[4]
            })

        if not formatted_reqs:
            logger.warning(f"No structured requirements found for project {project_id}.")
            
        # 3. Fetch Skill Instructions
        cursor.execute("SELECT skill_name, target_doc_type, instructions FROM skills WHERE id = %s", (skill_id,))
        skill_res = cursor.fetchone()
        cursor.close()
        conn.close()

        if not skill_res:
            return False, "Selected skill not found in database."
            
        skill_name, target_doc_type, instructions = skill_res
        
        # 3.5 Fetch Reference Document (if any)
        reference_context = ""
        if reference_document_id:
            cursor.execute("SELECT original_filename, full_markdown_content FROM documents WHERE doc_id = %s::uuid", (reference_document_id,))
            doc_res = cursor.fetchone()
            if doc_res:
                ref_filename, ref_content = doc_res
                reference_context = f"""
# Reference Document ({ref_filename})
You MUST use this document as the primary reference for your analysis and generation. Do not generate content outside the scope defined in this document.
{ref_content}
"""
            else:
                logger.warning(f"Reference document {reference_document_id} not found.")

        # 4. Call Gemini
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return False, "GEMINI_API_KEY is not configured."
            
        genai.configure(api_key=api_key)
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
        model = genai.GenerativeModel(model_name)
        
        prompt = f"""
You are an expert QA Automation Engineer, Business Analyst, and Technical Writer.
Your task is to generate a formal QA Document based on the provided System Knowledge, Skill Instructions, and Reference Documents.

# Target Document Information
- Document Name: {doc_name}
- Document Type: {doc_type}
- Project Code: {project_code}

# Framework & Instructions (Skill: {skill_name})
Please follow these instructions strictly to structure and generate the document:
{instructions}

{reference_context}

# System Knowledge (Phase 1 Structured Requirements)
Use the following system requirements to populate the document with accurate, relevant information.
If the knowledge is empty, try your best to create a generic template or deduce from the document name.
{json.dumps(formatted_reqs, ensure_ascii=False, indent=2)}

# Output Format
Generate the complete document in standard Markdown format. Use professional formatting, tables if necessary, and clear headings.
Do NOT wrap the entire response in ```markdown ... ``` blocks unless necessary, just output the raw markdown text directly.
"""
        logger.info(f"Generating document '{doc_name}' ({doc_type}) using skill '{skill_name}'...")
        resp = model.generate_content(prompt)
        doc_content = resp.text.strip()
        
        # Strip markdown code blocks if the model wrapped the whole thing
        if doc_content.startswith("```markdown"):
            doc_content = doc_content[11:]
        if doc_content.startswith("```"):
            doc_content = doc_content[3:]
        if doc_content.endswith("```"):
            doc_content = doc_content[:-3]
            
        return True, doc_content.strip()

    except Exception as e:
        logger.error(f"Error in create_qa_document: {e}", exc_info=True)
        return False, str(e)


def create_qa_document_async(gen_id: str, project_id: str, doc_type: str, doc_name: str, skill_id: int, reference_document_id=None):
    """
    Async background version of QA Document Creator that generates Excel.
    """
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Fetch Project Info
        cursor.execute("SELECT project_code FROM projects WHERE id = %s::uuid", (project_id,))
        project_res = cursor.fetchone()
        project_code = project_res[0] if project_res else "Unknown Project"

        # 2. Fetch Knowledge
        cursor.execute("""
            SELECT req_code, title, description, steps, expected_results
            FROM structured_requirements
            WHERE project_id = %s::uuid
        """, (project_id,))
        reqs = cursor.fetchall()
        
        formatted_reqs = []
        for req in reqs:
            formatted_reqs.append({
                "req_code": req[0],
                "title": req[1],
                "description": req[2],
                "steps": req[3],
                "expected_results": req[4]
            })
            
        # 3. Fetch Skill
        cursor.execute("SELECT skill_name, target_doc_type, instructions FROM skills WHERE id = %s", (skill_id,))
        skill_res = cursor.fetchone()
        if not skill_res:
            raise ValueError("Skill not found.")
        skill_name, target_doc_type, instructions = skill_res
        
        # 3.5 Fetch Reference
        reference_context = ""
        if reference_document_id:
            cursor.execute("SELECT original_filename, full_markdown_content FROM documents WHERE doc_id = %s::uuid", (reference_document_id,))
            doc_res = cursor.fetchone()
            if doc_res:
                ref_filename, ref_content = doc_res
                reference_context = f"""
# Reference Document ({ref_filename})
You MUST use this document as the primary reference for your analysis and generation. Do not generate content outside the scope defined in this document.
{ref_content}
"""
        
        # 4. Call Gemini
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            raise ValueError("GEMINI_API_KEY is not configured.")
            
        genai.configure(api_key=api_key)
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
        model = genai.GenerativeModel(model_name)
        
        if doc_type == "Test Case":
            prompt = f"""
You are an expert QA Automation Engineer, Business Analyst, and Technical Writer.
Your task is to generate a formal QA Document based on the provided System Knowledge, Skill Instructions, and Reference Documents.

# Target Document Information
- Document Name: {doc_name}
- Document Type: {doc_type}
- Project Code: {project_code}

# Framework & Instructions (Skill: {skill_name})
Please follow these instructions strictly to structure and generate the document:
{instructions}

{reference_context}

# System Knowledge (Phase 1 Structured Requirements)
{json.dumps(formatted_reqs, ensure_ascii=False, indent=2)}

# Output Format MUST BE JSON
You MUST generate the entire document as a strict JSON object with two keys: "metadata" and "test_cases".
Do NOT include any text outside the JSON.
Format:
{{
  "metadata": {{
    "project_name": "{project_code} (Admin)",
    "tester_name": "AI Agent",
    "module_function": "Determined from requirements"
  }},
  "test_cases": [
    {{
      "Test Case ID": "TC-001",
      "Test case Objective": "...",
      "Test Description / Procedure": "1. ...\\n2. ...",
      "Test Data": "...",
      "Expected Result": "...",
      "Actual Result": "...",
      "Result (Pass/Fail)": "PASS",
      "Req No.": "...",
      "Update by": "AI Agent"
    }}
  ]
}}
"""
        else:
            prompt = f"""
You are an expert QA Automation Engineer, Business Analyst, and Technical Writer.
Your task is to generate a formal QA Document based on the provided System Knowledge, Skill Instructions, and Reference Documents.

# Target Document Information
- Document Name: {doc_name}
- Document Type: {doc_type}
- Project Code: {project_code}

# Framework & Instructions (Skill: {skill_name})
Please follow these instructions strictly to structure and generate the document:
{instructions}

{reference_context}

# System Knowledge (Phase 1 Structured Requirements)
{json.dumps(formatted_reqs, ensure_ascii=False, indent=2)}

# Output Format MUST BE JSON ARRAY
You MUST generate the entire document as a strict JSON Array of Objects.
Do NOT include any text outside the JSON array.
Each object in the array represents a single row in the final Excel file.
The keys of the objects will become the column headers. Make sure all objects use the same keys.
"""

        logger.info(f"Generating Excel document async '{doc_name}' ({doc_type})...")
        resp = model.generate_content(prompt)
        doc_content = resp.text.strip()
        
        # Clean JSON
        if doc_content.startswith("```json"): doc_content = doc_content[7:]
        elif doc_content.startswith("```"): doc_content = doc_content[3:]
        if doc_content.endswith("```"): doc_content = doc_content[:-3]
        doc_content = doc_content.strip()
        
        try:
            data = json.loads(doc_content)
        except json.JSONDecodeError:
            raise ValueError("AI did not return a valid JSON format.")
            
        # 5. Save to Excel
        import uuid
        upload_dir = os.path.join(os.getcwd(), 'uploads', 'qa_generated')
        os.makedirs(upload_dir, exist_ok=True)
        file_name = f"{doc_name.replace(' ', '_')}_{uuid.uuid4().hex[:6]}.xlsx"
        file_path = os.path.join(upload_dir, file_name)

        if doc_type == "Test Case":
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Case"
            
            # Styles
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Row 1: Title
            ws.merge_cells('A1:I1')
            ws['A1'] = "Test Case Specification"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_align
            
            # Row 3-6: Metadata
            meta = data.get("metadata", {})
            import datetime
            today = datetime.datetime.now().strftime("%d/%m/%Y")
            
            metadata_map = [
                ("Project Name :", meta.get("project_name", project_code), "Create Date :", today),
                ("Project ID:", project_code, "Start Test Date :", today),
                ("Tester Name :", meta.get("tester_name", "AI Agent"), "Finish Test Date :", today),
                ("Project Release / Version :", "-", "Module / Function:", meta.get("module_function", ""))
            ]
            
            row_idx = 3
            for r_data in metadata_map:
                ws.cell(row=row_idx, column=2).value = r_data[0]
                ws.cell(row=row_idx, column=2).font = bold_font
                ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
                
                ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
                ws.cell(row=row_idx, column=3).value = r_data[1]
                
                ws.cell(row=row_idx, column=6).value = r_data[2]
                ws.cell(row=row_idx, column=6).font = bold_font
                ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
                
                ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
                ws.cell(row=row_idx, column=7).value = r_data[3]
                
                # Apply light blue background to metadata area
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).fill = header_fill
                row_idx += 1
                
            # Row 7: Functional Requirements
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            ws.cell(row=row_idx, column=1).value = "FUNCTIONAL REQUIREMENTS (Requirements No.) :"
            ws.cell(row=row_idx, column=1).font = bold_font
            ws.cell(row=row_idx, column=1).alignment = center_align
            
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=8)
            ws.cell(row=row_idx, column=6).value = "Site test UAT :"
            ws.cell(row=row_idx, column=6).font = bold_font
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
            
            # Row 8: Table Headers
            headers = ["Test Case ID", "Test case Objective", "Test Description / Procedure", "Test Data", 
                       "Expected Result", "Actual Result", "Result (Pass/Fail)", "Req No.", "Update by"]
            row_idx += 1
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                
            # Set Column Widths
            widths = [15, 30, 40, 20, 30, 25, 15, 10, 20]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                
            # Data Rows
            test_cases = data.get("test_cases", [])
            row_idx += 1
            for tc in test_cases:
                for col_idx, h in enumerate(headers, 1):
                    val = tc.get(h, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align if col_idx in [1, 7, 8, 9] else left_align
                    
                    if h == "Result (Pass/Fail)":
                        if str(val).upper() == "PASS":
                            cell.fill = pass_fill
                        elif str(val).upper() == "FAIL":
                            cell.fill = fail_fill
                row_idx += 1
                
            wb.save(file_path)

        else:
            import pandas as pd
            if not isinstance(data, list):
                data = [data]
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
        
        # 6. Update DB
        cursor.execute("UPDATE qa_generated_documents SET status = 'Completed', file_url = %s WHERE id = %s::uuid", (file_path, gen_id))
        conn.commit()
        logger.info(f"Successfully generated {file_path}")

    except Exception as e:
        logger.error(f"Error in create_qa_document_async: {e}", exc_info=True)
        if conn and cursor:
            try:
                cursor.execute("UPDATE qa_generated_documents SET status = 'Failed' WHERE id = %s::uuid", (gen_id,))
                conn.commit()
            except:
                pass
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    from dotenv import load_dotenv
    load_dotenv()
    # Test script execution
    print("Script loaded successfully.")
